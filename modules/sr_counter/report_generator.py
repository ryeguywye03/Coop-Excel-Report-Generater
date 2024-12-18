import pandas as pd
from modules.utils.logger_manager import LoggerManager
from PyQt6.QtWidgets import QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, QMessageBox, QFileDialog, QLabel
from PyQt6.QtCore import Qt
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side


class ReportGenerator:
    def __init__(self, progress_bar):
        self.progress_bar = progress_bar
        self.logger = LoggerManager()  # Initialize logger
        self.logger.log_info("ReportGenerator initialized")
        self.included_months = []  # Initialize as an instance variable
        self.grand_totals = {}  # Store grand totals here

    def generate_report(self, df: pd.DataFrame, selected_columns, start_date: datetime, end_date: datetime, start_time=None, end_time=None, sort_by=None, exclusions=None):
        """Generate the report by processing the DataFrame."""
        self.logger.log_info(f"Starting report generation from {start_date} to {end_date} with time frame {start_time} to {end_time}")
        self.logger.log_debug(f"Selected columns: {selected_columns}, Sort by: {sort_by}")

        try:
            # Filter dates and times
            df = self._filter_dates(df, start_date, end_date, start_time, end_time)
            self.logger.log_debug(f"Row count after date and time filtering: {len(df)}")

            # Ensure the 'Month' column is created
            if 'Created Date' in df.columns:
                df = df.copy()  # Avoid SettingWithCopyWarning by working with a copy
                df.loc[:, 'Month'] = df['Created Date'].dt.month
                self.logger.log_debug("Month column successfully added to the DataFrame.")
            else:
                raise KeyError("'Created Date' column is missing in the DataFrame.")

            # Apply exclusions
            df = self._apply_exclusions(df, exclusions)
            self.logger.log_debug(f"Row count after applying exclusions: {len(df)}")

            # Define the months included
            self.included_months = self._get_month_range(start_date, end_date)[0]  # only keep month names
            self.logger.log_debug(f"Included months for the report: {self.included_months}")

            # Process groups
            report_data, totals = self._process_groups(df, selected_columns, start_date, end_date)
            combined_report = self._create_report_dataframe(report_data, selected_columns)
            self.logger.log_debug(f"Row count after processing groups: {len(combined_report)}")

            # Apply sorting if needed
            if sort_by and sort_by in combined_report.columns:
                combined_report.sort_values(by=sort_by, inplace=True)
                self.logger.log_info(f"Sorted by {sort_by}")

            # Append totals row
            combined_report = self._append_totals(combined_report, totals, selected_columns)
            self.logger.log_debug(f"Row count after appending totals (if applied): {len(combined_report)}")

            self.logger.log_info("Report generation completed successfully.")
            return combined_report
        except KeyError as ke:
            self.logger.log_error(f"KeyError in report generation: {ke}")
            return None
        except Exception as e:
            self.logger.log_error(f"Error generating report: {e}")
            return None



    def _filter_dates(self, df, start_date, end_date, start_time=None, end_time=None):
        """
        Filter the DataFrame by the specified date range and optional time frame.
        """
        self.logger.log_debug("Filtering data by date range and time frame")

        # Ensure 'Created Date' is a valid datetime
        df['Created Date'] = pd.to_datetime(df['Created Date'], errors='coerce')
        df.dropna(subset=['Created Date'], inplace=True)

        # Filter by date range
        date_filtered_df = df[(df['Created Date'] >= start_date) & (df['Created Date'] <= end_date)]

        if start_time and end_time:
            # Filter by time frame
            date_filtered_df = date_filtered_df[
                (date_filtered_df['Created Date'].dt.time >= start_time) &
                (date_filtered_df['Created Date'].dt.time <= end_time)
            ]

        return date_filtered_df


    def _apply_exclusions(self, df, exclusions):
        """Apply exclusions to the DataFrame based on the provided exclusion criteria."""
        if exclusions:
            self.logger.log_debug(f"Applying exclusions: {exclusions}")
            for column, exclude_values in exclusions.items():
                if column in df.columns:
                    original_row_count = len(df)
                    df = df[~df[column].isin(exclude_values)]
                    excluded_count = original_row_count - len(df)
                    self.logger.log_debug(f"Excluded {excluded_count} rows from column '{column}' with values: {exclude_values}")
                else:
                    self.logger.log_warning(f"Column '{column}' not found in DataFrame, unable to apply exclusion for this column.")
        return df
    
    def _append_totals(self, df, totals, selected_columns):
        """Append a totals row to the DataFrame, ensuring no duplicates."""
        self.logger.log_debug("Checking for existing totals row before appending.")

        # Check if a totals row already exists
        if df['TOTAL'].eq(totals['TOTAL']).any():
            self.logger.log_warning("Totals row already exists; skipping append.")
            return df  # Return the original DataFrame without modification

        # Prepare totals row
        total_row = {col: '' for col in selected_columns}  # Empty for grouping columns
        total_row[selected_columns[0]] = "Totals"  # Set "Totals" title in the first grouping column
        total_row.update(totals)  # Add totals to the row

        # Append row
        self.logger.log_debug("Appending totals row to the report.")
        totals_df = pd.DataFrame([total_row])
        df = pd.concat([df, totals_df], ignore_index=True)
        self.logger.log_info("Totals row appended successfully.")

        return df


    def _get_month_range(self, start_date, end_date):
        # Return both month names and month numbers
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        month_numbers = list(range(start_date.month, end_date.month + 1))
        month_names = months[start_date.month - 1:end_date.month]
        return month_names, month_numbers


    def _process_groups(self, df, selected_columns, start_date, end_date):
        self.logger.log_debug("Processing groups")
        combined_data = []
        month_names, month_numbers = self._get_month_range(start_date, end_date)
        totals = {month: 0 for month in month_names}
        totals['TOTAL'] = 0

        grouped = df.groupby(selected_columns)
        for group_values, group_data in grouped:
            monthly_counts = []
            
            # Calculate counts for only the included months (now using month numbers)
            for month_number in month_numbers:
                count = group_data[group_data['Month'] == month_number].shape[0]
                monthly_counts.append(count)

            # Calculate total for each row
            total_count = sum(monthly_counts)
            row_data = {col: val for col, val in zip(selected_columns, group_values)}
            row_data.update({**dict(zip(month_names, monthly_counts)), 'TOTAL': total_count})
            combined_data.append(row_data)

            # Update grand totals
            for month_name, count in zip(month_names, monthly_counts):
                totals[month_name] += count
            totals['TOTAL'] += total_count

        return combined_data, totals




    def _create_report_dataframe(self, data, selected_columns):
        self.logger.log_debug("Creating DataFrame for report")
        df = pd.DataFrame(data)
        return df[selected_columns + self.included_months + ['TOTAL']]


    def show_report_preview(self, preview_df, start_time=None, end_time=None):
        """
        Display the report preview with optional time frame details.
        """
        self.logger.log_info("Displaying report preview")
        try:
            dialog = QDialog()
            dialog.setWindowTitle("Report Preview")
            dialog.setWindowFlag(Qt.WindowType.WindowMaximizeButtonHint)
            layout = QVBoxLayout(dialog)

            # Add time frame details at the top if applicable
            if start_time and end_time:
                time_frame_label = QLabel(f"Time Frame: {start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}")
                time_frame_label.setStyleSheet("font-weight: bold; margin-bottom: 10px;")
                layout.addWidget(time_frame_label)

            table = QTableWidget()
            table.setRowCount(preview_df.shape[0])
            table.setColumnCount(preview_df.shape[1])
            table.setHorizontalHeaderLabels(preview_df.columns)

            for i in range(preview_df.shape[0]):
                for j in range(preview_df.shape[1]):
                    table.setItem(i, j, QTableWidgetItem(str(preview_df.iloc[i, j])))

            layout.addWidget(table)
            dialog.setLayout(layout)
            dialog.resize(800, 400)
            table.resizeColumnsToContents()

            dialog.exec()
        except Exception as e:
            self.logger.log_error(f"Error displaying preview: {e}")


    def _create_preview_dialog(self, df):
        dialog = QDialog()
        dialog.setWindowTitle("Report Preview")
        dialog.setWindowFlag(Qt.WindowType.WindowMaximizeButtonHint)
        layout = QVBoxLayout(dialog)

        table = QTableWidget()
        table.setRowCount(df.shape[0])
        table.setColumnCount(df.shape[1])
        table.setHorizontalHeaderLabels(df.columns)

        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                table.setItem(i, j, QTableWidgetItem(str(df.iloc[i, j])))

        layout.addWidget(table)
        dialog.setLayout(layout)
        dialog.resize(800, 400)
        table.resizeColumnsToContents()

        return dialog

    def save_report(self, report_df, start_date, end_date, start_time=None, end_time=None):
        """Save the report to an Excel file with optional time frame details."""
        self.logger.log_info("Saving report to Excel")

        try:
            # Drop empty rows
            report_df.dropna(how='all', inplace=True)

            # Prompt for save location
            file_path = self._prompt_save_file()
            if not file_path:
                self.logger.log_info("Save canceled by user.")
                # QMessageBox.warning(None, "Save Canceled", "The report was not saved because the save operation was canceled.")
                return None

            # Save to Excel
            self._save_to_excel(report_df, file_path, start_time=start_time, end_time=end_time)
            # QMessageBox.information(self, "Success", f"Report saved at {file_path}")
            return file_path
        except Exception as e:
            self.logger.log_error(f"Error saving report: {e}")
            QMessageBox.critical(None, "Error", "Failed to save report.")
            return None



    def _prompt_save_file(self):
        current_time = datetime.now()
        default_filename = current_time.strftime("sr_count_report-%m-%d-%Y-%H-%M-%S.xlsx")
        file_path, _ = QFileDialog.getSaveFileName(
            None,
            "Save Report",
            os.path.join(os.path.expanduser("~/Desktop"), default_filename),
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            self.logger.log_warning("No file path selected; save operation canceled.")
            return None

        self.logger.log_info(f"File path selected: {file_path}")
        return file_path



    def _save_to_excel(self, df, file_path, start_time=None, end_time=None):
        """
        Save the DataFrame to an Excel file with optional time frame details.

        :param df: The DataFrame to save.
        :param file_path: The path to save the Excel file.
        :param start_time: Start time for the time frame (if enabled).
        :param end_time: End time for the time frame (if enabled).
        """
        self.logger.log_debug("Writing DataFrame to Excel with time frame (if enabled).")
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Report'

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            time_frame_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            total_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            total_font = Font(bold=True)

            current_row = 1

            # Add time frame details if applicable
            if start_time and end_time:
                time_frame_text = f"Time Frame: {start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}"
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(df.columns))
                time_frame_cell = worksheet.cell(row=current_row, column=1, value=time_frame_text)
                time_frame_cell.font = Font(bold=True)
                time_frame_cell.fill = time_frame_fill
                current_row += 1  # Move to the next row for headers

            # Write headers
            for col_num, column_title in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=current_row, column=col_num, value=column_title)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = cell_border

            # Write data rows
            for row_num, row_data in enumerate(df.itertuples(index=False), start=current_row + 1):
                is_total_row = row_data[0] == "Totals"
                for col_num, value in enumerate(row_data, start=1):
                    cell = worksheet.cell(row=row_num, column=col_num, value=value)
                    if is_total_row:
                        cell.font = total_font
                        cell.fill = total_fill
                    cell.border = cell_border

            # Autofit column widths safely
            for col_idx, column_cells in enumerate(worksheet.iter_cols(min_row=current_row, max_row=current_row, max_col=len(df.columns)), start=1):
                column_letter = column_cells[0].column_letter
                max_length = max(len(str(cell.value or "")) for cell in worksheet[column_letter]) + 2
                worksheet.column_dimensions[column_letter].width = max_length

            workbook.save(file_path)
            self.logger.log_debug(f"Excel file saved successfully to {file_path}.")
        except Exception as e:
            self.logger.log_error(f"Failed to save Excel file: {e}")


