import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from PyQt6.QtWidgets import QFileDialog
from modules.utils.logger_manager import LoggerManager


class SRFormatter:
    def __init__(self, logger=None):
        self.logger = logger if logger else LoggerManager()

    def format_sr_data(self, input_file):
        """Format the SR data and save it to an Excel file."""
        try:
            # Load the input file
            df = pd.read_excel(input_file)

            # Process the data
            df = self._process_data(df)

            # Prompt for save location
            current_time = datetime.now()
            default_filename = current_time.strftime("formatted_sr_report-%m-%d-%Y-%H-%M-%S.xlsx")
            file_path, _ = QFileDialog.getSaveFileName(
                None,
                "Save Formatted SR Data",
                default_filename,
                "Excel Files (*.xlsx)"
            )
            if not file_path:
                self.logger.log_info("Save operation canceled by the user.")
                return

            # Save the formatted data
            self._save_to_excel(df, file_path)
            self.logger.log_info(f"Formatted data saved to {file_path}")

            return file_path
        except Exception as e:
            self.logger.log_error(f"Error formatting SR data: {e}")
            raise

    def preview_sr_data(self, input_file, rows=None):
        """Preview the SR data after formatting, showing the specified number of rows or all rows if not specified."""
        try:
            # Load the file
            df = pd.read_excel(input_file)

            # Process the data
            df = self._process_data(df)

            # If rows is None, preview the entire file
            preview_df = df if rows is None else df.head(rows)

            # Return the preview of the data
            self.logger.log_info("Generating preview for SR data.")
            return preview_df
        except Exception as e:
            self.logger.log_error(f"Error generating preview: {e}")
            raise


    def _is_already_processed(self, row, sr_column_index):
        """
        Check if a row is already processed based on SR # and the column to its right.
        The column to the right of SR # is dynamically identified.
        """
        if sr_column_index + 1 < len(row):
            return pd.notnull(row.iloc[sr_column_index + 1])
        return False

    def _process_data(self, df):
        """Process the DataFrame by splitting and formatting the 'SR #' column."""
        if 'SR #' not in df.columns:
            raise KeyError("'SR #' column is missing in the input file.")

        self.logger.log_info("Processing SR data.")

        # Get the index of the 'SR #' column
        sr_column_index = df.columns.get_loc('SR #')

        # Iterate through each row and update inline if necessary
        for index, row in df.iterrows():
            if pd.notnull(row['SR #']):
                # Skip processing if the row is already processed
                if self._is_already_processed(row, sr_column_index):
                    continue

                # Process the SR # column
                sr_parts = str(row['SR #']).split(' ', 1)
                sr_number = self._add_leading_zeros(sr_parts[0]) if len(sr_parts) > 0 else None
                description = sr_parts[1] if len(sr_parts) > 1 else ''

                # Update the SR # column
                df.at[index, 'SR #'] = sr_number

                # Update the column to the right of SR #
                right_column_name = df.columns[sr_column_index + 1]
                df.at[index, right_column_name] = description

        # Rename the column to the right of SR # to "Description"
        right_column_name = df.columns[sr_column_index + 1]
        if 'Unnamed' in right_column_name:  # Check if it's an unnamed column
            df.rename(columns={right_column_name: 'Description'}, inplace=True)

        # Reorder columns to ensure SR # and Description are first
        description_column = 'Description'
        columns_order = ['SR #', description_column] + [col for col in df.columns if col not in ['SR #', description_column]]
        return df[columns_order]



    @staticmethod
    def _add_leading_zeros(sr_number):
        """Add leading zeros to the numeric part of the SR number."""
        try:
            prefix, number = sr_number.split('-')
            number = number.zfill(8)
            return f"{prefix}-{number}"
        except Exception:
            return sr_number  # Return as-is if parsing fails

    def _save_to_excel(self, df, file_path, start_time=None, end_time=None):
        """
        Save the DataFrame to an Excel file with optional time frame details.

        :param df: The DataFrame to save.
        :param file_path: The path to save the Excel file.
        :param start_time: Start time for the time frame (if enabled).
        :param end_time: End time for the time frame (if enabled).
        """
        self.logger.log_debug("Writing DataFrame to Excel with time frame (if enabled).")
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Formatted SR Data"

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            time_frame_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

            current_row = 1

            # Add time frame details if applicable
            if start_time and end_time:
                time_frame_text = f"Time Frame: {start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}"
                worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(df.columns))
                time_frame_cell = worksheet.cell(row=current_row, column=1, value=time_frame_text)
                time_frame_cell.font = Font(bold=True)
                time_frame_cell.fill = time_frame_fill
                current_row += 1  # Move to the next row for headers

            # Write headers
            for col_num, column_title in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=current_row, column=col_num, value=column_title)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = cell_border

            # Write data rows
            for row_num, row_data in enumerate(df.itertuples(index=False), start=current_row + 1):
                for col_num, value in enumerate(row_data, start=1):
                    cell = worksheet.cell(row=row_num, column=col_num, value=value)
                    cell.border = cell_border

            # Autofit column widths
            for col_idx, column_cells in enumerate(worksheet.iter_cols(min_row=1, max_row=worksheet.max_row, max_col=len(df.columns)), start=1):
                column_letter = column_cells[0].column_letter
                max_length = max(len(str(cell.value or "")) for cell in column_cells) + 2
                worksheet.column_dimensions[column_letter].width = max_length

            workbook.save(file_path)
            self.logger.log_debug(f"Excel file saved successfully to {file_path}")
        except Exception as e:
            self.logger.log_error(f"Failed to save Excel file: {e}")
